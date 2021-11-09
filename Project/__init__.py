from flask import Flask, request, abort, render_template, redirect, url_for
import requests
import json
from Project.Config import *
#from uncleengineer import thaistock
# import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from itertools import islice

app = Flask(__name__)




@app.route('/')
# def hello():
#     return 'hello world Somchai51', 200
def show_Excel_openpyxl():

    # aa = "Hello"
    aa = ["Hello", "Every one"]
    # aa = ("tuple1", "2", 33)
    # aa = ["Hello", ("tuple1", "2", 33), ["lis2", 10, 233]]

    # data = tuple_to_str(aa)
    # data = list_to_str(aa)
    # ss = "แควอ้อม ม.8"
    ss = "บาง"
    # data = openpy_630422(ss)
    data = openpy_641008(ss)
    return render_template('excel.html', data=data)

@app.route('/excelopenpyxl')



@app.route('/webhook', methods=['POST', 'GET'])
def webhook():
    if request.method == 'POST':
        payload = request.json
        print(payload)
        print()
        Reply_token = payload['events'][0]['replyToken']
        print(Reply_token)
        print()
        message = payload['events'][0]['message']['text']
        print('message:')
        print(message)
        print(type(message))
        print()
        dest = payload['destination']
        print(dest)
        print()
        R0 = payload['events'][0]
        print(R0)
        print()
        # use for elif "btc"
        aa = "Hello"
        # aa = ["Hello", "Every one"]
        # aa = ("tuple1", "2", 33)
        # aa = ["Hello", ("tuple1", "2", 33), ["lis2", 10, 233]]

        # user for else:
        bb = {
            "type": "text",
            "text": "Hello"
        }
        cc = {
            "type": "text",
            "text": "(2565)นายสมชาย น้อยเอี่ยม work from home ที่ เอกชัยแมนชั่น ถ.เอกชัย ต.แม่กลอง อ.เมือง สมุทรสงคราม"
        }

        dd = {
            "type": "flex",
            "altText": "Flex Message",
            "contents": {
                "type": "bubble",
                "direction": "ltr",
                "header": {
                    "type": "box",
                    "layout": "vertical",
                    "contents": [
                        {
                            "type": "text",
                            "text": "Header",
                            "align": "center"
                        }
                    ]
                },
                "hero": {
                    "type": "image",
                    "url": "https://developers.line.biz/assets/images/services/bot-designer-icon.png",
                    "size": "full",
                    "aspectRatio": "1.51:1",
                    "aspectMode": "fit"
                },
                "body": {
                    "type": "box",
                    "layout": "vertical",
                    "contents": [
                        {
                            "type": "text",
                            "text": "Hello World",
                            "align": "center"
                        }
                    ]
                },
                "footer": {
                    "type": "box",
                    "layout": "horizontal",
                    "contents": [
                        {
                            "type": "button",
                            "action": {
                                "type": "uri",
                                "label": "Button",
                                "uri": "https://linecorp.com"
                            }
                        }
                    ]
                }
            }
        }

        # data = demo_reader_jsonfile_todict()
        # data = dd

        if 'hun' in message:
            ITD = thaistock('ITD')
            Reply_messasge = 'ราคาหุ้น อิตาเลียนไทย ขณะนี้ : {}'.format(ITD)
            print('Re:' + Reply_messasge)
            ReplyMessage(Reply_token, Reply_messasge, Channel_access_token)

        elif "btc" in message:
            # Reply_messasge = 'ราคา BITCOIN ขณะนี้ : {}'.format(GET_BTC_PRICE())
            aa = list_to_str(aa)
            # aa = tuple_to_str(aa)
            Reply_messasge = 'Btc:\n{}'.format(aa)
            print('Re:' + Reply_messasge)
            ReplyMessage(Reply_token, Reply_messasge, Channel_access_token)
        else:
            # data = openpy_630422(message)
            data = openpy_641008(message)
            if data == 1:
                # data = demo_reader_jsonfile_todict()
                data = cc

            Reply_messasge = data
            print('Data in webhook Function:')
            print(Reply_messasge)
            print('data_type:')
            print(type(Reply_messasge))
            print()
            ReplyMessage(Reply_token, Reply_messasge, Channel_access_token)

        return request.json, 200

    elif request.method == 'GET':
        return 'this is method GET from Webhook!!!', 200
    else:
        abort(400)


# ในlist มีหลาย tuple ใน tuple มีหลายสติง ก็จะแตกเป็นสติงให้ด้วย


def ReplyMessage(Reply_token, TextMessage, Line_Acees_Token):
    LINE_API = 'https://api.line.me/v2/bot/message/reply'

    Authorization = 'Bearer {}'.format(Line_Acees_Token)  # ที่ยาวๆ
    print('Data(TextMessage) in ReplyMeassage Function')
    print(TextMessage)
    print('data_type:')
    print(type(TextMessage))
    print()
    print(Authorization)
    print()
    headers = {
        'Content-Type': 'application/json; charset=UTF-8',
        'Authorization': Authorization
    }

    if type(TextMessage) == type(""):
        print('str1')
        data = {
            "replyToken": Reply_token,
            "messages": [{
                "type": "text",
                "text": TextMessage
            }]
            # "messages": [TextMessage]
        }
    elif type(TextMessage) == type({}):
        print('dict2')
        data = {
            "replyToken": Reply_token,
            # "messages": [{
            #     "type": "text",
            #     "text": TextMessage
            # }]
            "messages": [TextMessage]
        }

    data = json.dumps(data)  # dump dict >> Json Object
    print('data:')
    print(data)
    print('data_type:')
    print(type(data))
    return requests.post(LINE_API, headers=headers, data=data), 200






def str_to_list(s):
    # sep = ","
    print('------------in str_to_list ฟังก์ชัน----------')
    print(s)
    print(type(s))
    sep = ","  # แยกด้วย sep
    l = list(s.split(sep))
    print(l)
    print(type(l))
    print(len(l))
    print('-------------exit str_to_list ฟังก์ชัน----------')
    print()
    return l


def list_to_tuple_to_str_to_list(l):  # ในlist มี tuple ก็จะแตกเป็นสติงให้ด้วย
    print('------------in list_to_tuple_to_str_to_list ฟังก์ชัน----------')
    print(l)
    print(type(l))
    new_list = []
    for s in l:  # วนลูปแปลงแต่ละสมาชิก
        if type(s) == type(()):
            s = tuple_to_str(s)  # ใช้ฟังก์ชัน tuple_to_str(t) ทำให้เป็น str
            print()
            print("S")
            print(s)
            print(type(s))
            print()

        else:
            print(s)
            print(type(s))
            print()
        new_list.append(s)  # รวมสมาชิกที่เป็นสติงแล้วเข้าใน list ใหม่

    print("new_list")
    print(new_list)
    print(type(new_list))
    print()
    print('-----------list_to_tuple_to_str_to_list ฟังก์ชัน----------')
    print()
    return new_list


def list_to_tuple(l):  # ในlist มี tuple อยู่  only 1 เท่านั้น
    print('------------in list_to_tuple ฟังก์ชัน----------')
    print(l)
    print(type(l))
    for t in l:  # วนลูปแปลงแต่ละสมาชิก
        print()
        print("tuple")
        print(t)
        print(type(t))
        print()

    print('-----------exit-list_to_tuple ฟังก์ชัน----------')
    print()
    return t


def tuple_to_str(t):  # t-tuple
    # sep = ","
    print('------------in tuple_to_str ฟังก์ชัน----------')
    print(type(t))

    sep = ","  # รวมคั่นด้วย sep
    newss = ""
    for i in t:
        i = str(i)  # transfome to string type
        print(i)
        print(type(i))
        newss = newss+i+sep
    newss = newss[0:-1]
    print()
    print(newss)
    print(type(newss))
    print(len(newss))
    print('-------------exit tuple_to_str ฟังก์ชัน----------')
    print()
    return newss


def list_to_str(l):     # not work # use l[0] l[1] instead
    print('------------in list_to_str ฟังก์ชัน----------')
    # sep = ","
    sep = "\n"
    print(l)
    print(type(l))
    s = sep.join(l)
    print(s)
    print(type(s))
    print('------------Exit list_to_str ฟังก์ชัน----------')
    print()
    return s


def openpy(wanted):
    print(wanted)
    print(type(wanted))
    print()
    wb = load_workbook('./Project/assign_numbering.xlsx')
    # Get a sheet by name
    sheet = wb.get_sheet_by_name('6301')
    data = sheet.values
    data_islice = [r[5:10] for r in islice(data, 1)]
    print('----IN----islice-------')
    print(data_islice)
    print(type(data_islice))
    print()
    print('----Exit----islice-------')
    print()
    data_islice = list_to_tuple(data_islice)
    data_islice = tuple_to_str(data_islice)
    data_islice = str_to_list(data_islice)

    # Convert your data to a list(inside = tuple)
    data = list(data)
    print('data after turn to list')
    # print(data)
    print(len(data))
    print(type(data))
    print()

    idx = [r[5] for r in data]
    # print(idx)

    print(len(idx))
    print(type(idx))
    print()
    print("*"*30)
    if wanted in idx:
        for i in range(0, len(idx)):
            if idx[i] == wanted:
                data_net = data[i][5:]
                print(type(data_net))
        print('---End1---')

        data = tuple_to_str(data_net)
        data = str_to_list(data)
        print(type(data))
    else:
        data = 1

    newss = ""
    sep = ","
    if data != 1:
        for i in range(0, len(data)):
            d1 = data_islice[i]+" : " + data[i]
            newss = newss+d1+sep
        newss = newss[0:-1]
        print(newss)
        print(type(newss))
        data = newss

    return data
###############################################################
############################---new--###############################

def openpy_641008(wanted):
    print(wanted)
    print(type(wanted))
    print()
    wb = load_workbook('./Project/assign_numbering.xlsx')
    # Get a sheet by name
    sheet = wb.get_sheet_by_name('6301')
    data = sheet.values
    print(data)
    # print(len(data))
    print(type(data))
    data_islice = [r[5:10] for r in islice(data, 1)]
    print('///////IN----islice-------')
    print(data_islice)
    print(type(data_islice))
    print()
    print('///////Exit----islice-------')
    print()
    data_islice = list_to_tuple(data_islice)
    print()
    data_islice = tuple_to_str(data_islice)
    print()
    # data_islice = str_to_list(data_islice)
    print()
    print('//////Exit----islice All Tranform-------')
    print()

    # Convert your data to a list(inside = tuple)
    data = list(data)
    print('data after turn to list')
    print(data)
    print(len(data))
    print(type(data))
    print()

    # idx = [r[5] for r in data]
    # print(idx)

    # print(len(idx))
    # print(type(idx))
    # print()
    print("*"*30)
    data_net1 = []
    data_net2 = []
    data_net3 = []
    data_net4 = []
    data_net5 = []
    for i in range(0, len(data)):  # last not included
        if data[i][5].find(wanted) != -1:
            data_net1.append(data[i][5])
            data_net2.append(data[i][6])
            data_net3.append(data[i][7])
            data_net4.append(data[i][8])
            data_net5.append(data[i][9])

        print('***************** End data_net Loop ', i+1, ' = ', data[i][5],'***********************')
        print()
        print(data_net1)
        print(type(data_net1))
        print()
    print('LEN of Data_net1:')
    print(len(data_net1))
    if len(data_net1) > 0:
        print()
        # data = list_to_tuple_to_string(data_net) #important function
        d_som=""
        for i in range(0, len(data_net1)):
            d1="ชุมสาย "+str(data_net1[i])+"\r\n"
            d2="Active = "+str(data_net2[i])+"\r\n"
            d3="Inactive = "+str(data_net3[i])+"\r\n"
            d4="ว่าง = "+str(data_net4[i])+"\r\n"
            d5="รวมทั้งหมด = "+str(data_net5[i])+"\r\n"

            be="***ข้อมูลที่ "+str((i+1))+"***\r\n"

            d_som=d_som+be+d1+d2+d3+d4+d5+"\r\n"
        # for r in data_net:
        #     dd="ชุมสาย--"+r+"\r\n"
        #     d_som=d_som+dd
        # data = data_islice+"\n"+d_som
        data = "ข้อมูล["+wanted+"]มีทั้งหมด "+str(len(data_net1))+" รายการ\r\n" + d_som
        print('Summary Data in openpy_630422 Function:')
        print(data)
        print(type(data))
        print()
    else:
        data = 1

    return data
###############################################################




def openpy_630422(wanted):
    print(wanted)
    print(type(wanted))
    print()
    wb = load_workbook('./Project/assign_numbering.xlsx')
    # Get a sheet by name
    sheet = wb.get_sheet_by_name('6301')
    data = sheet.values
    print(data)
    # print(len(data))
    print(type(data))
    data_islice = [r[5:10] for r in islice(data, 1)]
    print('///////IN----islice-------')
    print(data_islice)
    print(type(data_islice))
    print()
    print('///////Exit----islice-------')
    print()
    data_islice = list_to_tuple(data_islice)
    print()
    data_islice = tuple_to_str(data_islice)
    print()
    # data_islice = str_to_list(data_islice)
    print()
    print('//////Exit----islice All Tranform-------')
    print()

    # Convert your data to a list(inside = tuple)
    data = list(data)
    print('data after turn to list')
    print(data)
    print(len(data))
    print(type(data))
    print()

    idx = [r[5] for r in data]
    print(idx)

    print(len(idx))
    print(type(idx))
    print()
    print("*"*30)
    data_net = []

    for i in range(0, len(idx)):  # last not included
        if idx[i].find(wanted) != -1:
            data_net.append(data[i][5:])

        print('***************** End data_net Loop ', i+1, ' = ', idx[i])
        print()
        print(data_net)
        print(type(data_net))
        print()
    print('LEN of Data_net:')
    print(len(data_net))
    if len(data_net) > 0:
        print()
        data = list_to_tuple_to_string(data_net) #important function
        data = data_islice+"\n"+data
        print('Summary Data in openpy_630422 Function:')
        print(data)
        print(type(data))
        print()
    else:
        data = 1

    return data
###############################################################
def list_to_tuple_to_string(l):
    print('------------in list_to_tuple_to_string ฟังก์ชัน----------')
    print(l)
    print(type(l))
    sep = ("\n")
    data = ""
    for t in l:  # วนลูปแปลงแต่ละสมาชิก
        print()
        print("tuple")
        print(t)
        print(type(t))
        print()
        data = data+tuple_to_str(t)+sep
    print('-----------exit-list_to_tuple_to_string ฟังก์ชัน----------')
    data = data[0:-1]
    print()
    return data
############################################################################
def openpy2(wanted):
    print(wanted)
    print(type(wanted))
    print()
    wb = load_workbook('./Project/assign_numbering.xlsx')

    # Get a sheet by name
    sheet = wb.get_sheet_by_name('6301')

    data = sheet.values

    # Convert your data to a list(inside = tuple)
    data = list(data)
    print('data after turn to list')
    # print(data)
    print(len(data))
    print(type(data))
    print()
    # Read in the data at index 0 for the indices
    # idx = [r for r in data]  # MAKE New List Commulate  in rows at idx
    # idx = [r[0] for r in data]    #MAKE New List Commulate in specific Col0 at idx
    # idx = [r[1] for r in data]    #MAKE New List Commulate in specific Col1 at idx
    # MAKE New List Commulate in specific Col5 at idx
    idx = [r[5] for r in data]
    # print(idx)
    print(len(idx))
    print(type(idx))

    print("*"*30)
    if wanted in idx:
        for i in range(0, len(idx)):
            if idx[i] == wanted:
                data_net = data[i][5:]
                print(type(data_net))
        print('---End1---')

        data = tuple_to_str(data_net)
        data_net = str_to_list((data))
        # print(type(data))
        data = {
            "type": "flex",
            "altText": "this is a flex message",
            "contents":
            {
                "type": "bubble",
                "body": {
                    "type": "box",
                    "layout": "vertical",
                    "contents": [
                        {
                            "type": "text",
                            "text": ":: --- ::",
                            "weight": "bold",
                            "align": "center"
                        },
                        {
                            "type": "text",
                            "text": "ชุมสาย : " + data_net[0],
                            "weight": "bold",
                            "color": "#0000ff"
                        },
                        {
                            "type": "text",
                            "text": "Active         : " + data_net[1],
                            "color": "#32CD32"
                        },
                        {
                            "type": "text",
                            "text": "Inactive      : " + data_net[2],
                            "color": "#ff0000"
                        },
                        {
                            "type": "text",
                            "text": "(ว่าง)           : " + data_net[3]
                        },
                        {
                            "type": "text",
                            "text": "รวมทั้งหมด : " + data_net[4],
                            "weight": "bold"
                        }
                    ]
                }
            }
        }

    else:
        data = 1

    return data


def GET_BTC_PRICE():
    data = requests.get('https://bx.in.th/api/')
    BTC_PRICE = data.text.split('BTC')[1].split('last_price":')[
        1].split(',"volume_24hours')[0]
    return BTC_PRICE


def demo_reader_unicode():  # Not use
    # with open("C:\\Users\\Winet\\Desktop\\data_line.txt", mode="r", encoding="utf8") as f:
    # with open(r"C:\Users\Winet\Desktop\data_line.txt", mode="r", encoding="utf8") as f:
    with open(r"C:/Users/somchain/Desktop/data_line.txt", mode="r", encoding="utf8") as f:
        print("=========in def demo_reader_unicode()=========")
        print('f.name = ' + f.name)
        print('f.mode = ' + f.mode)
        data = f.read()
        print('data:')
        print(data)
        print(type(data))
        print("=========out def demo_reader_unicode()=========")
        return data


def demo_reader_jsonfile_todict():  # use
    # with open("C:\\Users\\Winet\\Desktop\\data_line.txt", mode="r", encoding="utf8") as f:
    # with open(r"C:\Users\Winet\Desktop\data_line.txt", mode="r", encoding="utf8") as f:
    with open(r"C:/Users/somchain/Desktop/data.json", mode="r", encoding="utf8") as f:
        print("=========in def demo_reader_jsonfile_todict()=========")
        print('f.name = ' + f.name)
        print('f.mode = ' + f.mode)
        data = json.load(f)
        print('data:')
        print(data)
        print(type(data))
        print("=========out def demo_reader_jsonfile_todict()=========")
        return data






    # return 200


# @app.route('/webhook', methods=['POST', 'GET'])
# def webhook():
#     if request.method == 'POST':
#         print(request.json)
#         return request.json, 200
#     elif request.method == 'GET':
#         return 'this is method GET!!', 200
#     else:
#         abort(400)


# @app.route('/')
# def hello():
#     return 'Hello World',200


# @app.route('/webhook', methods=['GET'])
# def hello2():
#     return 'This is wEBHOOK',200
